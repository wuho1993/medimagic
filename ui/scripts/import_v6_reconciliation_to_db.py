#!/usr/bin/env python3
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


UI_ROOT = Path(__file__).resolve().parents[1]
RECONCILIATION_FILE = Path('/Users/joecheung/Desktop/V6_員工資料對照_record.xlsx')
EMPLOYEE_MASTER_FILE = Path('/Users/joecheung/Desktop/員工資料.xlsx')
OUTPUT_SQL = UI_ROOT / 'supabase' / 'manual' / '20260504_import_v6_reconciliation.sql'
PAYROLL_MONTH_START = '2026-03-01'

MATCHED_SHEET = 'V6_對到員工資料'
UNMATCHED_SHEET = 'V6_員工資料未有或未確認'

BRANCH_MAP = {
    'TM': 'TM',
    'TMA': 'TM',
    'TaiWai': 'TAIWAI',
    'TAIWAI': 'TAIWAI',
    'Tai Wai': 'TAIWAI',
    'Office': 'OFFICE',
    'OFFICE': 'OFFICE',
    'MKTOP': 'MKTOP',
    'MK TOP': 'MKTOP',
    'MK': 'MKTOP',
    'MKCY': 'MKCY',
    'MOS': 'MOS',
    'Mos': 'MOS',
    'TW': 'TW',
    'CWB': 'CWB',
    'CWA': 'CWB',
    '醫療': 'OFFICE',
}

POSITION_MAP = {
    'Senior Beautician': 'SENIOR_BEAUTICIAN',
    'Beautician': 'BEAUTICIAN',
    'Beautician(PT)': 'BEAUTICIAN',
    'Massagist': 'MASSAGIST',
    'Consultant': 'CONSULTANT',
    'Sales Manager': 'SALES_MANAGER',
    'Manager': 'MANAGER',
    'Assistant Shop Manager': 'ASSISTANT_SHOP_MANAGER',
    'Senior Shop Manager': 'SENIOR_SHOP_MANAGER',
    'Senior Sales Manager': 'SENIOR_SALES_MANAGER',
    'Account Manager': 'ACCOUNT_MANAGER',
    'Operation Manager': 'OPERATION_MANAGER',
    'Assistant Operation Manager': 'ASSISTANT_OPERATION_MANAGER',
    'Receptionist': 'RECEPTIONIST',
    'RECEPTIONIST': 'RECEPTIONIST',
    'Trainer': 'TRAINER',
    'Designer': 'DESIGNER',
    'Masketing Executive': 'MARKETING_EXECUTIVE',
    'Marketing Executive': 'MARKETING_EXECUTIVE',
    'Office Clerk': 'OFFICE_CLERK',
    'Telesales': 'TELESALES',
    'Janitor': 'JANITOR',
    '治療師': 'BEAUTICIAN',
    '養生師': 'MASSAGIST',
    '店長': 'SHOP_MANAGER',
    '經理': 'MANAGER',
    '接待員': 'RECEPTIONIST',
    '老闆': 'BOSS',
    '營運總監': 'OPERATION_DIRECTOR',
    '營運經理': 'OPERATION_MANAGER',
    '美容動畫設計師': 'DESIGNER',
    '行銷主管': 'MARKETING_EXECUTIVE',
    '培訓老師': 'TRAINER',
    '客戶服務主管': 'CUSTOMER_SERVICE_MANAGER',
    '會計經理': 'ACCOUNT_MANAGER',
    '顧問': 'CONSULTANT',
    '會計文員': 'ACCOUNTING_CLERK',
    '文員': 'OFFICE_CLERK',
    '電話銷售員': 'TELESALES',
    '清潔員': 'JANITOR',
    '跟針姑娘': 'NURSE',
    '醫生護士': 'NURSE',
    '曾醫師': 'DOCTOR',
}

POSITION_SEEDS = {
    'MARKETING_EXECUTIVE': ('行銷主管', 'Marketing Executive'),
    'RECEPTIONIST': ('接待員', 'Receptionist'),
    'TRAINER': ('培訓老師', 'Trainer'),
    'DESIGNER': ('設計師', 'Designer'),
    'CONSULTANT': ('顧問', 'Consultant'),
    'BEAUTICIAN': ('美容師', 'Beautician'),
    'MASSAGIST': ('養生師', 'Massagist'),
    'JANITOR': ('清潔員', 'Janitor'),
    'TELESALES': ('電話銷售員', 'Telesales'),
    'OFFICE_CLERK': ('文員', 'Office Clerk'),
    'ACCOUNTING_CLERK': ('會計文員', 'Accounting Clerk'),
    'SHOP_MANAGER': ('店長', 'Shop Manager'),
    'MANAGER': ('經理', 'Manager'),
    'NURSE': ('護士', 'Nurse'),
    'DOCTOR': ('醫師', 'Doctor'),
    'BOSS': ('老闆', 'Boss'),
    'ACCOUNT_MANAGER': ('會計經理', 'Account Manager'),
    'OPERATION_MANAGER': ('營運經理', 'Operation Manager'),
    'OPERATION_DIRECTOR': ('營運總監', 'Operation Director'),
    'CUSTOMER_SERVICE_MANAGER': ('客戶服務主管', 'Customer Service Manager'),
    'SENIOR_BEAUTICIAN': ('高級美容師', 'Senior Beautician'),
    'SALES_MANAGER': ('銷售經理', 'Sales Manager'),
    'SENIOR_SALES_MANAGER': ('高級銷售經理', 'Senior Sales Manager'),
    'SENIOR_SHOP_MANAGER': ('高級店長', 'Senior Shop Manager'),
    'ASSISTANT_SHOP_MANAGER': ('副店長', 'Assistant Shop Manager'),
    'ASSISTANT_OPERATION_MANAGER': ('助理營運經理', 'Assistant Operation Manager'),
}


@dataclass
class MasterRow:
    sheet: str
    row_number: int
    branch: str | None
    position: str | None
    alias: str | None
    phone: str | None
    name_zh: str | None
    name_en: str | None
    dob: str | None
    identity_number: str | None
    hire_date: str | None
    address: str | None


def safe_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r'\s+', ' ', text)


def parse_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    match = re.match(r'^(\d{4})年(\d{1,2})月(\d{1,2})日$', text)
    if match:
        year, month, day = map(int, match.groups())
        return date(year, month, day).isoformat()

    match = re.match(r'^(\d{1,2})/(\d{1,2})/?/?(\d{4})$', text)
    if match:
        day, month, year = map(int, match.groups())
        return date(year, month, day).isoformat()

    match = re.match(r'^(\d{4})-(\d{2})-(\d{2})', text)
    if match:
        return f'{match.group(1)}-{match.group(2)}-{match.group(3)}'

    return None


def contains_cjk(text: str | None) -> bool:
    return bool(text and re.search(r'[\u3400-\u9fff]', text))


def normalize_branch(value: str | None) -> str | None:
    if not value:
        return None
    cleaned = value.strip()
    if cleaned in BRANCH_MAP:
        return BRANCH_MAP[cleaned]
    upper = cleaned.upper()
    if upper in BRANCH_MAP:
        return BRANCH_MAP[upper]
    if contains_cjk(cleaned):
        return None
    if re.fullmatch(r'[A-Za-z0-9/_ -]+', cleaned):
        return upper
    return None


def normalize_company(value: str | None) -> str:
    if not value:
        return 'ASA'
    normalized = value.strip().upper()
    if normalized in {'ASA', 'A'}:
        return 'ASA'
    if normalized == 'ASAS':
        return 'ASAS'
    return 'ASA'


def normalize_position(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    if text in POSITION_MAP:
        return POSITION_MAP[text]
    for raw, code in POSITION_MAP.items():
        if text.lower() == raw.lower():
            return code
    return None


def parse_number(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(float(value)).rstrip('0').rstrip('.') if '.' in str(float(value)) else str(int(value))
    text = str(value).strip()
    if not text or text in {'N/A', '-', '/hr', '/day'}:
        return None
    match = re.search(r'-?\d+(?:\.\d+)?', text.replace(',', ''))
    if not match:
        return None
    number = match.group(0)
    return number.rstrip('0').rstrip('.') if '.' in number else number


def sql_text(value: str | None) -> str:
    if value is None or value == '':
        return 'null'
    escaped = value.replace("'", "''")
    return f"'{escaped}'"


def sql_required_text(value: str | None) -> str:
    escaped = (value or '').replace("'", "''")
    return f"'{escaped}'"


def sql_date(value: str | None) -> str:
    if not value:
        return 'null'
    return f"date '{value}'"


def sql_number(value: str | None) -> str:
    return value if value not in {None, ''} else 'null'


def build_merge_text(target: str, field_name: str) -> str:
    return (
        f"case "
        f"when excluded.{field_name} is null or excluded.{field_name} = '' then {target}.{field_name} "
        f"when {target}.{field_name} is null or {target}.{field_name} = '' then excluded.{field_name} "
        f"when position(excluded.{field_name} in {target}.{field_name}) > 0 then {target}.{field_name} "
        f"else {target}.{field_name} || E'\\n' || excluded.{field_name} end"
    )


def load_master_rows() -> dict[tuple[str, int], MasterRow]:
    workbook = load_workbook(EMPLOYEE_MASTER_FILE, data_only=True)
    result: dict[tuple[str, int], MasterRow] = {}

    for sheet_name in ('現職', '離職'):
        sheet = workbook[sheet_name]
        for row_number in range(2, sheet.max_row + 1):
            result[(sheet_name, row_number)] = MasterRow(
                sheet=sheet_name,
                row_number=row_number,
                branch=safe_text(sheet.cell(row=row_number, column=2).value),
                position=safe_text(sheet.cell(row=row_number, column=3).value),
                alias=safe_text(sheet.cell(row=row_number, column=4).value),
                phone=safe_text(sheet.cell(row=row_number, column=6).value),
                name_zh=safe_text(sheet.cell(row=row_number, column=7).value),
                name_en=safe_text(sheet.cell(row=row_number, column=8).value),
                dob=parse_date(sheet.cell(row=row_number, column=9).value),
                identity_number=safe_text(sheet.cell(row=row_number, column=10).value),
                hire_date=parse_date(sheet.cell(row=row_number, column=11).value),
                address=safe_text(sheet.cell(row=row_number, column=12).value),
            )

    return result


def normalize_name_pair(name_zh: str | None, name_en: str | None, fallback_name: str | None) -> tuple[str, str]:
    if name_zh and name_en:
        return name_zh, name_en
    if name_zh and not name_en:
        return name_zh, name_zh
    if name_en and not name_zh:
        return name_en, name_en
    fallback = fallback_name or 'Unknown'
    return (fallback, fallback) if contains_cjk(fallback) else (fallback, fallback)


def build_employee_code(v6_staff_no: str | None, reason: str | None, v6_row: int) -> tuple[str, str | None]:
    if v6_staff_no and v6_staff_no != 'N/A' and not reason_contains_temporary_code(reason):
        return v6_staff_no, None
    original = v6_staff_no if v6_staff_no and v6_staff_no != 'N/A' else None
    return f'V6TMP202603R{v6_row:03d}', original


def reason_contains_temporary_code(reason: str | None) -> bool:
    if not reason:
        return False
    return '重複' in reason or '未能確認' in reason


def build_employee_note(v6_row: int, reason: str | None, original_staff_no: str | None, placeholder_hire_date: bool, matched: bool) -> str:
    parts = [f'V6 匯入 2026-03；來源 row {v6_row}。']
    if matched:
        parts.append('已對到員工資料。')
    if reason:
        parts.append(reason.rstrip('。') + '。')
    if original_staff_no:
        parts.append(f'原始 V6 staff no：{original_staff_no}。')
    if placeholder_hire_date:
        parts.append(f'入職日缺失，暫以 {PAYROLL_MONTH_START} 佔位。')
    return ''.join(parts)


def build_salary_remark(v6_row: int, matched: bool, reason: str | None) -> str:
    parts = [f'V6 2026-03 匯入；row {v6_row}']
    if matched:
        parts.append('；已對到員工資料')
    if reason:
        parts.append(f'；{reason}')
    return ''.join(parts)


def load_sheet_rows(sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(RECONCILIATION_FILE, data_only=True)
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    rows: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {str(headers[index]): values[index] for index in range(len(headers))}
        rows.append(row)
    return rows


def to_record(raw: dict[str, object], master_rows: dict[tuple[str, int], MasterRow], matched: bool) -> dict[str, object]:
    v6_row = int(raw['V6 Row'])
    v6_staff_no = safe_text(raw.get('V6 Staff No'))
    v6_name = safe_text(raw.get('V6 Name'))
    v6_branch = safe_text(raw.get('V6 Branch'))
    v6_company = normalize_company(safe_text(raw.get('V6 Company')))
    v6_note = safe_text(raw.get('V6 Note'))
    reason = None if matched else safe_text(raw.get('Reason'))

    master_row = None
    if matched:
        employee_sheet = safe_text(raw.get('Employee Sheet'))
        employee_row = raw.get('Employee Row')
        if employee_sheet and employee_row:
            master_row = master_rows.get((employee_sheet, int(employee_row)))

    employee_code, original_staff_no = build_employee_code(v6_staff_no, reason, v6_row)
    name_zh, name_en = normalize_name_pair(
        master_row.name_zh if master_row else safe_text(raw.get('Name ZH')),
        master_row.name_en if master_row else safe_text(raw.get('Name EN')),
        v6_name,
    )

    branch_code = normalize_branch(
        (master_row.branch if master_row else safe_text(raw.get('Employee Branch'))) or v6_branch,
    )
    position_code = normalize_position(
        (master_row.position if master_row else safe_text(raw.get('Position'))),
    )

    alias = (master_row.alias if master_row else safe_text(raw.get('Alias'))) or v6_name
    phone = master_row.phone if master_row else None
    address = master_row.address if master_row else None
    identity_number = master_row.identity_number if master_row and master_row.identity_number else ''
    date_of_birth = master_row.dob if master_row else None
    hire_date = master_row.hire_date if master_row and master_row.hire_date else PAYROLL_MONTH_START
    placeholder_hire_date = not (master_row and master_row.hire_date)

    salary_type = 'monthly'
    if any('/hr' in str(raw.get(key) or '').lower() for key in ('Basic', 'Atten', 'Meeting', 'Booking', 'Transport')):
        salary_type = 'hourly'
    elif any('/day' in str(raw.get(key) or '').lower() for key in ('Basic', 'Atten', 'Meeting', 'Booking', 'Transport')):
        salary_type = 'daily'

    employment_type = '全職'
    if reason and '兼職' in reason:
        employment_type = '兼職'
    elif salary_type in {'hourly', 'daily'}:
        employment_type = '兼職'

    return {
        'employee_code': employee_code,
        'original_staff_no': original_staff_no,
        'v6_row': v6_row,
        'name_zh': name_zh,
        'name_en': name_en,
        'alias': alias,
        'company_type': v6_company,
        'branch_code': branch_code,
        'position_code': position_code,
        'phone': phone,
        'address': address,
        'identity_number': identity_number,
        'date_of_birth': date_of_birth,
        'hire_date': hire_date,
        'placeholder_hire_date': placeholder_hire_date,
        'employment_type': employment_type,
        'employment_status': 'active',
        'base_salary': parse_number(raw.get('Basic')),
        'attendance_bonus_amount': parse_number(raw.get('Atten')),
        'briefing_bonus': parse_number(raw.get('Meeting')),
        'booking_bonus': parse_number(raw.get('Booking')),
        'transport_allowance': parse_number(raw.get('Transport')),
        'salary_type': salary_type,
        'employee_notes': build_employee_note(v6_row, reason, original_staff_no, placeholder_hire_date, matched),
        'salary_remarks': build_salary_remark(v6_row, matched, reason),
        'commission_notes': v6_note,
        'matched': matched,
        'reason': reason,
    }


def build_position_seed_sql(position_codes: set[str]) -> str:
    values = []
    for code in sorted(position_codes):
        if code not in POSITION_SEEDS:
            continue
        name_zh, name_en = POSITION_SEEDS[code]
        values.append(f"({sql_text(code)}, {sql_text(name_zh)}, {sql_text(name_en)})")
    if not values:
        return ''
    joined = ',\n  '.join(values)
    return (
        '-- Ensure referenced positions exist\n'
        'insert into public.positions (code, name_zh, name_en)\n'
        f'values\n  {joined}\n'
        'on conflict (code) do update set\n'
        '  name_zh = excluded.name_zh,\n'
        '  name_en = excluded.name_en;\n'
    )


def build_employee_sql(record: dict[str, object]) -> str:
    return f"""
insert into public.employees (
  employee_code,
  name_zh,
  name_en,
  alias,
  gender,
  identity_type,
  identity_number,
  date_of_birth,
  address,
  phone,
  company_type,
  company_id,
  branch_id,
  branch_code,
  employment_type,
  employment_status,
  position_id,
  hire_date,
  notes
)
values (
  {sql_text(record['employee_code'])},
  {sql_text(record['name_zh'])},
  {sql_text(record['name_en'])},
  {sql_text(record['alias'])},
  'other'::public.employee_gender,
  'hkid'::public.employee_identity_type,
    {sql_required_text(record['identity_number'])},
  {sql_date(record['date_of_birth'])},
  {sql_text(record['address'])},
  {sql_text(record['phone'])},
  {sql_text(record['company_type'])}::public.employee_company_type,
  (select id from public.companies where code = {sql_text(record['company_type'])} limit 1),
  (select id from public.branches where code = {sql_text(record['branch_code'])} limit 1),
  {sql_text(record['branch_code'])},
  {sql_text(record['employment_type'])}::public.employee_employment_type,
  {sql_text(record['employment_status'])}::public.employee_status,
  (select id from public.positions where code = {sql_text(record['position_code'])} limit 1),
  {sql_date(record['hire_date'])},
  {sql_text(record['employee_notes'])}
)
on conflict (employee_code) do update set
  name_zh = coalesce(nullif(excluded.name_zh, ''), employees.name_zh),
  name_en = coalesce(nullif(excluded.name_en, ''), employees.name_en),
  alias = coalesce(nullif(excluded.alias, ''), employees.alias),
  identity_number = case when excluded.identity_number is null then employees.identity_number else excluded.identity_number end,
  date_of_birth = coalesce(excluded.date_of_birth, employees.date_of_birth),
  address = coalesce(nullif(excluded.address, ''), employees.address),
  phone = coalesce(nullif(excluded.phone, ''), employees.phone),
  company_type = coalesce(excluded.company_type, employees.company_type),
  company_id = coalesce(excluded.company_id, employees.company_id),
  branch_id = coalesce(excluded.branch_id, employees.branch_id),
  branch_code = coalesce(nullif(excluded.branch_code, ''), employees.branch_code),
  employment_type = coalesce(excluded.employment_type, employees.employment_type),
  employment_status = coalesce(excluded.employment_status, employees.employment_status),
  position_id = coalesce(excluded.position_id, employees.position_id),
  hire_date = coalesce(excluded.hire_date, employees.hire_date),
  notes = {build_merge_text('employees', 'notes')},
  updated_at = timezone('utc', now());
""".strip()


def build_salary_sql(record: dict[str, object]) -> str:
    attendance_amount = sql_number(record['attendance_bonus_amount'])
    attendance_enabled = 'true' if record['attendance_bonus_amount'] not in {None, '0'} else 'false'
    return f"""
insert into public.employee_salary_profiles (
  employee_id,
  salary_type,
  base_salary,
  effective_from,
  remarks,
  attendance_bonus_enabled,
  attendance_bonus_amount,
  transport_allowance,
  briefing_bonus,
  booking_bonus,
  commission_notes
)
values (
  (select id from public.employees where employee_code = {sql_text(record['employee_code'])} limit 1),
  {sql_text(record['salary_type'])}::public.employee_salary_type,
  {sql_number(record['base_salary'])},
  {sql_date(record['hire_date'])},
  {sql_text(record['salary_remarks'])},
  {attendance_enabled},
  {attendance_amount},
  {sql_number(record['transport_allowance'])},
  {sql_number(record['briefing_bonus'])},
  {sql_number(record['booking_bonus'])},
  {sql_text(record['commission_notes'])}
)
on conflict (employee_id) do update set
  salary_type = coalesce(excluded.salary_type, employee_salary_profiles.salary_type),
  base_salary = coalesce(excluded.base_salary, employee_salary_profiles.base_salary),
  effective_from = coalesce(excluded.effective_from, employee_salary_profiles.effective_from),
  attendance_bonus_enabled = coalesce(excluded.attendance_bonus_enabled, employee_salary_profiles.attendance_bonus_enabled),
  attendance_bonus_amount = coalesce(excluded.attendance_bonus_amount, employee_salary_profiles.attendance_bonus_amount),
  transport_allowance = coalesce(excluded.transport_allowance, employee_salary_profiles.transport_allowance),
  briefing_bonus = coalesce(excluded.briefing_bonus, employee_salary_profiles.briefing_bonus),
  booking_bonus = coalesce(excluded.booking_bonus, employee_salary_profiles.booking_bonus),
  remarks = {build_merge_text('employee_salary_profiles', 'remarks')},
  commission_notes = {build_merge_text('employee_salary_profiles', 'commission_notes')},
  updated_at = timezone('utc', now());
""".strip()


def main() -> None:
    master_rows = load_master_rows()
    matched_rows = [to_record(row, master_rows, True) for row in load_sheet_rows(MATCHED_SHEET)]
    unmatched_rows = [to_record(row, master_rows, False) for row in load_sheet_rows(UNMATCHED_SHEET)]
    all_rows = matched_rows + unmatched_rows

    position_codes = {record['position_code'] for record in all_rows if record['position_code']}
    statements = ['begin;', '']

    position_seed_sql = build_position_seed_sql({code for code in position_codes if isinstance(code, str)})
    if position_seed_sql:
        statements.append(position_seed_sql)
        statements.append('')

    statements.append('-- Employee upserts from V6 reconciliation')
    for record in all_rows:
        statements.append(build_employee_sql(record))
        statements.append('')

    statements.append('-- Salary profile upserts from V6 reconciliation')
    for record in all_rows:
        statements.append(build_salary_sql(record))
        statements.append('')

    statements.append('commit;')

    OUTPUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL.write_text('\n'.join(statements), encoding='utf-8')

    temp_codes = sum(1 for record in all_rows if str(record['employee_code']).startswith('V6TMP202603'))
    print(f'WROTE_SQL={OUTPUT_SQL}')
    print(f'TOTAL_ROWS={len(all_rows)}')
    print(f'MATCHED_ROWS={len(matched_rows)}')
    print(f'UNMATCHED_ROWS={len(unmatched_rows)}')
    print(f'TEMP_EMPLOYEE_CODES={temp_codes}')


if __name__ == '__main__':
    main()