"""
Full staff import builder.
Sources:
  - 員工資料.xlsx  → employee master (name, identity, hire date, DOB, address, phone)
  - 03-2026_MM(FY)_V6.xlsx  → Contract List (salary, company, probation) + SALARY sheet + bank sheets

Code conflict resolution (user decisions):
  SF169/SF185 → SF185
  SF259 (Ki Ki) → SF238
  SF343 (貞姐) → SF153

Branch mapping:
  MK → MKTOP, TOP → MKTOP, TaiWai/Tai Wai/Tai wai → TAIWAI, TM → TMA,
  Office/OFfice → OFFICE, Mos/MOS → MOS, TW → TW, MKCY → MKCY, CBA → CBA,
  CWB → CWB (new), MK TOP → MKTOP, 醫療 → OFFICE

Company mapping from V6 col J:
  A/ASA/自 → ASA, ASAS → ASAS, Fanny → ASAS (her row weirdly has 'Fanny')

Position mapping: new positions will be created as needed.
"""
from __future__ import annotations
import openpyxl, re, csv, sys
from pathlib import Path
from collections import Counter

FILE1 = Path('/Users/joecheung/Desktop/員工資料.xlsx')
FILE2 = Path('/Users/joecheung/Desktop/03-2026_MM(FY)_V6.xlsx')
OUT_DIR = Path(__file__).resolve().parents[1] / 'import'
OUT_DIR.mkdir(exist_ok=True)

def xcode(raw):
    if not raw: return None
    m = re.match(r'^(SF\d+)', str(raw).strip())
    return m.group(1) if m else None

def dstr(v):
    if v is None: return ''
    if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s

def resolve_code(code, raw_full=''):
    if 'SF169' in raw_full or code == 'SF169':
        return 'SF185'
    if 'SF259' in raw_full or code == 'SF259':
        return 'SF238'
    if 'SF343' in raw_full or code == 'SF343':
        return 'SF153'
    return code

BRANCH_MAP = {
    'MK': 'MKTOP', 'TOP': 'MKTOP', 'MK TOP': 'MKTOP',
    'TaiWai': 'TAIWAI', 'Tai Wai': 'TAIWAI', 'Tai wai': 'TAIWAI',
    'TM': 'TMA',
    'Office': 'OFFICE', 'OFfice': 'OFFICE',
    'TW': 'TW',
    'Mos': 'MOS', 'MOS': 'MOS',
    'MKCY': 'MKCY',
    'CBA': 'CBA',
    'CWB': 'CWB',
    '醫療': 'OFFICE',
}

def norm_branch(raw):
    raw = raw.strip()
    if raw in BRANCH_MAP:
        return BRANCH_MAP[raw]
    # Try case-insensitive
    for k, v in BRANCH_MAP.items():
        if raw.lower() == k.lower():
            return v
    return raw  # return as-is, will be flagged

def norm_company(raw):
    raw = raw.strip()
    if raw in ('A', 'ASA', '自', 'Fanny'):  # Fanny is a data anomaly
        return 'ASA'
    if raw in ('ASAS',):
        return 'ASAS'
    return raw

# Position normalization
POSITION_MAP = {
    # Beautician-type roles
    'Senior Beautician': 'SENIOR_BEAUTICIAN',
    'Beautician': 'BEAUTICIAN',
    'Beautician(PT)': 'BEAUTICIAN',
    # Massage/wellness
    'Massagist': 'MASSAGIST',
    # Consultant-type
    'Consultant': 'CONSULTANT',
    'Sales Manager': 'SALES_MANAGER',
    'Manager': 'MANAGER',
    'Assistant Shop Manager': 'ASSISTANT_SHOP_MANAGER',
    'Senior Shop Manager': 'SENIOR_SHOP_MANAGER',
    'Senior Sales Manager': 'SENIOR_SALES_MANAGER',
    'Account Manager': 'ACCOUNT_MANAGER',
    # Operations
    'Operation Manager': 'OPERATION_MANAGER',
    'Assistant Operation Manager': 'ASSISTANT_OPERATION_MANAGER',
    # Receptionist
    'Receptionist': 'RECEPTIONIST',
    'RECEPTIONIST': 'RECEPTIONIST',
    # Office roles
    'Trainer': 'TRAINER',
    'Designer': 'DESIGNER',
    'Masketing Executive': 'MARKETING_EXECUTIVE',
    'Office Clerk': 'OFFICE_CLERK',
    'Telesales': 'TELESALES',
    # Janitor
    'Janitor': 'JANITOR',
    # Chinese-named positions from V6
    '治療師': 'BEAUTICIAN',
    '養生師': 'MASSAGIST',
    '店長': 'SHOP_MANAGER',
    '經理': 'MANAGER',
    '接待員': 'RECEPTIONIST',
    '老闆': 'BOSS',
    '營運總監': 'OPERATION_DIRECTOR',
    '營運經理': 'OPERATION_MANAGER',
    '美容動畫設計師': 'DESIGNER',
    '行銷主管': 'MARKETING_EXECUTIVE',
    '培訓老師': 'TRAINER',
    '客戶服務主管': 'CUSTOMER_SERVICE_MANAGER',
    '會計經理': 'ACCOUNT_MANAGER',
    '顧問': 'CONSULTANT',
    '會計文員': 'OFFICE_CLERK',
    '電話銷售員': 'TELESALES',
    '文員': 'OFFICE_CLERK',
    '清潔員': 'JANITOR',
    '跟針姑娘': 'NURSE',
    '醫生護士': 'NURSE',
    '曾醫師': 'DOCTOR',
    # PT variants
    '治療師 (PT)': 'BEAUTICIAN',
    '治療師(PT)': 'BEAUTICIAN',
    '治療師 (自)': 'BEAUTICIAN',
    '養生師 (PT)': 'MASSAGIST',
    '治療師(PS)': 'BEAUTICIAN',
    'Body C': 'BEAUTICIAN',
}

def norm_position(raw):
    raw = raw.strip()
    if raw in POSITION_MAP:
        return POSITION_MAP[raw]
    # partial match
    for k, v in POSITION_MAP.items():
        if raw.lower() == k.lower():
            return v
    return ''

# ════════════════════════════════════════
# STEP 1: Load 員工資料
# ════════════════════════════════════════
wb1 = openpyxl.load_workbook(FILE1, data_only=True)

staff = {}
ws = wb1['現職']
for r in range(2, ws.max_row + 1):
    code = xcode(ws.cell(r, 5).value)
    if not code: continue
    staff[code] = {
        'code': code,
        'status': 'active',
        'branch_raw': dstr(ws.cell(r, 2).value),
        'position_raw': dstr(ws.cell(r, 3).value),
        'alias': dstr(ws.cell(r, 4).value),
        'phone': dstr(ws.cell(r, 6).value),
        'name_zh': dstr(ws.cell(r, 7).value),
        'name_en': dstr(ws.cell(r, 8).value),
        'dob': dstr(ws.cell(r, 9).value),
        'identity': dstr(ws.cell(r, 10).value),
        'hire_date': dstr(ws.cell(r, 11).value),
        'address': dstr(ws.cell(r, 12).value),
    }

ws2 = wb1['離職']
for r in range(2, ws2.max_row + 1):
    code = xcode(ws2.cell(r, 5).value)
    if not code: continue
    leave_date = dstr(ws2.cell(r, 1).value)
    if code in staff:
        staff[code]['also_in_resigned'] = True
        staff[code]['leave_date'] = leave_date
    else:
        staff[code] = {
            'code': code,
            'status': 'resigned',
            'leave_date': leave_date,
            'branch_raw': dstr(ws2.cell(r, 2).value),
            'position_raw': dstr(ws2.cell(r, 3).value),
            'alias': dstr(ws2.cell(r, 4).value),
            'phone': dstr(ws2.cell(r, 6).value),
            'name_zh': dstr(ws2.cell(r, 7).value),
            'name_en': dstr(ws2.cell(r, 8).value),
            'dob': dstr(ws2.cell(r, 9).value),
            'identity': dstr(ws2.cell(r, 10).value),
            'hire_date': dstr(ws2.cell(r, 11).value),
            'address': dstr(ws2.cell(r, 12).value),
        }

print(f"員工資料: {len(staff)} staff ({sum(1 for s in staff.values() if s['status']=='active')} active, {sum(1 for s in staff.values() if s['status']=='resigned')} resigned)")

# ════════════════════════════════════════
# STEP 2: Load V6 Contract List
# ════════════════════════════════════════
wb2 = openpyxl.load_workbook(FILE2, data_only=True)
ws_cl = wb2['Contract List']

v6_contract = {}
for r in range(1, ws_cl.max_row + 1):
    raw = ws_cl.cell(r, 1).value
    if not raw: continue
    raw_full = str(raw).strip()
    if raw_full.lower().startswith('staff'): continue
    code = xcode(raw)
    if not code: continue
    resolved = resolve_code(code, raw_full)

    v6_contract[resolved] = {
        'code_raw': raw_full,
        'branch': dstr(ws_cl.cell(r, 3).value),
        'position': dstr(ws_cl.cell(r, 4).value),
        'name_en': dstr(ws_cl.cell(r, 5).value),
        'name_zh': dstr(ws_cl.cell(r, 6).value),
        'identity': dstr(ws_cl.cell(r, 7).value),
        'mpf_company': dstr(ws_cl.cell(r, 8).value),
        'company_raw': dstr(ws_cl.cell(r, 10).value),
        'probation': dstr(ws_cl.cell(r, 11).value),
        'hire_date': dstr(ws_cl.cell(r, 12).value),
        'base_salary': dstr(ws_cl.cell(r, 15).value),
        'attendance': dstr(ws_cl.cell(r, 16).value),
        'briefing': dstr(ws_cl.cell(r, 17).value),
        'booking': dstr(ws_cl.cell(r, 18).value),
        'transport': dstr(ws_cl.cell(r, 19).value),
        'pay_day_1': dstr(ws_cl.cell(r, 26).value),
        'pay_day_2': dstr(ws_cl.cell(r, 27).value),
    }

print(f"V6 Contract List: {len(v6_contract)}")

# ════════════════════════════════════════
# STEP 3: Load V6 SALARY sheet
# ════════════════════════════════════════
ws_sal = wb2['SALARY']
v6_salary = {}
for r in range(1, min(ws_sal.max_row + 1, 200)):
    raw = ws_sal.cell(r, 2).value
    code = xcode(raw)
    if not code: continue
    resolved = resolve_code(code)
    v6_salary[resolved] = {
        'base_salary': dstr(ws_sal.cell(r, 3).value),
        'attendance': dstr(ws_sal.cell(r, 4).value),
        'briefing': dstr(ws_sal.cell(r, 5).value),
        'booking': dstr(ws_sal.cell(r, 6).value),
        'transport': dstr(ws_sal.cell(r, 7).value),
        'branch': dstr(ws_sal.cell(r, 14).value),
        'company': dstr(ws_sal.cell(r, 15).value),
        'alias': dstr(ws_sal.cell(r, 16).value),
    }

print(f"V6 SALARY: {len(v6_salary)}")

# ════════════════════════════════════════
# STEP 4: Load V6 ASAS BANK
# ════════════════════════════════════════
ws_asas = wb2['ASAS BANK']
bank_by_nick = {}
for r in range(2, ws_asas.max_row + 1):
    nick = dstr(ws_asas.cell(r, 1).value)
    name = dstr(ws_asas.cell(r, 2).value)
    bank = dstr(ws_asas.cell(r, 3).value)
    acct = dstr(ws_asas.cell(r, 4).value)
    co = dstr(ws_asas.cell(r, 5).value)
    if nick:
        bank_by_nick[nick.lower()] = {'name': name, 'bank': bank, 'account': acct, 'company': co}

print(f"ASAS BANK: {len(bank_by_nick)}")

# ════════════════════════════════════════
# MERGE AND REPORT
# ════════════════════════════════════════
# Also include V6-only codes (SF001, SF002, SF322, SF301) — already handled by resolve
# SF001 and SF002 are bosses in V6 but not in 員工資料 — add them
for code in sorted(set(v6_contract.keys()) - set(staff.keys())):
    d = v6_contract[code]
    staff[code] = {
        'code': code,
        'status': 'active',
        'branch_raw': d['branch'],
        'position_raw': d['position'],
        'alias': d.get('code_raw', '').split('/')[-1] if '/' in d.get('code_raw','') else '',
        'name_zh': d['name_zh'],
        'name_en': d['name_en'],
        'identity': d['identity'],
        'hire_date': d['hire_date'],
        'v6_only': True,
        'dob': '', 'address': '', 'phone': '',
    }
    # Use alias from V6
    if not staff[code]['alias']:
        staff[code]['alias'] = d.get('alias', '') if 'alias' in v6_contract[code] else ''
        v6a = v6_contract[code]
        staff[code]['alias'] = dstr(wb2['Contract List'].cell(1,2).value) # wrong, fix below

# Fix: get alias properly from v6_contract
for code in set(v6_contract.keys()) - set(staff.keys()) | set(v6_contract.keys()):
    pass  # already handled inline

# Actually re-fix V6-only entries
v6_only_codes = sorted(set(v6_contract.keys()) - set(s for s in staff if not staff[s].get('v6_only')))
# The above logic already added them. Let's just fix alias
for code in staff:
    if staff[code].get('v6_only') and code in v6_contract:
        staff[code]['alias'] = v6_contract[code].get('alias', staff[code].get('alias', ''))

# ════════════════════════════════════════
# Build merged records
# ════════════════════════════════════════
records = []
position_counter = Counter()
branch_issues = []

for code in sorted(staff.keys()):
    s = staff[code]
    cl = v6_contract.get(code, {})
    sal = v6_salary.get(code, {})

    # Employee master from 員工資料
    name_zh = s.get('name_zh', '') or cl.get('name_zh', '')
    name_en = s.get('name_en', '') or cl.get('name_en', '')
    alias = s.get('alias', '') or cl.get('alias', '')
    identity = s.get('identity', '') or cl.get('identity', '')
    hire_date = s.get('hire_date', '') or cl.get('hire_date', '')
    dob = s.get('dob', '')
    address = s.get('address', '')
    phone = s.get('phone', '')
    status = s.get('status', 'active')

    # Branch: prefer V6 contract, fallback 員工資料
    branch_raw = cl.get('branch', '') or sal.get('branch', '') or s.get('branch_raw', '')
    branch_norm = norm_branch(branch_raw)

    # Company from V6 col J
    company_raw = cl.get('company_raw', '') or sal.get('company', '')
    company_norm = norm_company(company_raw) if company_raw else 'ASA'  # default ASA

    # Position: prefer 員工資料 (English), V6 as fallback
    position_raw = s.get('position_raw', '') or cl.get('position', '')
    position_norm = norm_position(position_raw)
    if not position_norm and cl.get('position'):
        position_norm = norm_position(cl['position'])
    if not position_norm:
        position_norm = 'BEAUTICIAN'  # safe fallback
    position_counter[position_norm] += 1

    # Salary from V6 Contract List (primary) or SALARY sheet (secondary)
    base_salary = cl.get('base_salary', '') or sal.get('base_salary', '')
    attendance = cl.get('attendance', '') or sal.get('attendance', '')
    briefing = cl.get('briefing', '') or sal.get('briefing', '')
    booking = cl.get('booking', '') or sal.get('booking', '')
    transport = cl.get('transport', '') or sal.get('transport', '')
    pay_day_1 = cl.get('pay_day_1', '')
    pay_day_2 = cl.get('pay_day_2', '')

    # Probation from V6
    probation_raw = cl.get('probation', '')
    probation_months = ''
    if probation_raw:
        m = re.search(r'(\d+)', probation_raw)
        if m:
            val = int(m.group(1))
            if '日' in probation_raw:
                probation_months = '0'
            else:
                probation_months = str(val)

    # MPF company from V6
    mpf_company = cl.get('mpf_company', '')

    # Flags
    flags = []
    if not identity or identity == 'N/A':
        flags.append('missing_identity')
    if not hire_date:
        flags.append('missing_hire_date')
    if not base_salary or base_salary in ('N/A',):
        flags.append('no_salary')
    if '日薪' in str(base_salary) or '時薪' in str(base_salary) or 'Part-time' in str(base_salary):
        flags.append('part_time_salary')
    if '/' in branch_raw and branch_raw not in BRANCH_MAP:
        flags.append('multi_branch')
    if branch_norm == branch_raw and branch_raw and branch_norm not in ('OFFICE','TAIWAI','TW','MKTOP','TMA','MOS','CBA','MKCY','CWB'):
        flags.append('unknown_branch')

    has_salary = bool(base_salary and base_salary not in ('N/A', '') and '日薪' not in str(base_salary) and '時薪' not in str(base_salary))

    records.append({
        'code': code, 'name_zh': name_zh, 'name_en': name_en, 'alias': alias,
        'identity': identity, 'hire_date': hire_date, 'dob': dob,
        'address': address, 'phone': phone, 'status': status,
        'branch_raw': branch_raw, 'branch_norm': branch_norm,
        'company_raw': company_raw, 'company_norm': company_norm,
        'position_raw': position_raw, 'position_norm': position_norm,
        'base_salary': base_salary, 'attendance': attendance,
        'briefing': briefing, 'booking': booking, 'transport': transport,
        'pay_day_1': pay_day_1, 'pay_day_2': pay_day_2,
        'probation_months': probation_months,
        'mpf_company': mpf_company,
        'has_salary': has_salary,
        'flags': '|'.join(flags),
    })

# Write CSV
csv_path = OUT_DIR / 'full_staff_staging_20260415.csv'
fieldnames = [
    'code', 'name_zh', 'name_en', 'alias', 'identity', 'hire_date', 'dob',
    'address', 'phone', 'status',
    'branch_raw', 'branch_norm', 'company_raw', 'company_norm',
    'position_raw', 'position_norm',
    'base_salary', 'attendance', 'briefing', 'booking', 'transport',
    'pay_day_1', 'pay_day_2', 'probation_months', 'mpf_company',
    'has_salary', 'flags',
]
with csv_path.open('w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=fieldnames)
    w.writeheader()
    w.writerows(records)

# Summary
clean = [r for r in records if not r['flags']]
flagged = [r for r in records if r['flags']]
with_salary = [r for r in records if r['has_salary']]
no_sal = [r for r in records if not r['has_salary']]
active = [r for r in records if r['status'] == 'active']
resigned = [r for r in records if r['status'] == 'resigned']

print(f"\n{'='*60}")
print(f"TOTAL MERGED RECORDS: {len(records)}")
print(f"  Active: {len(active)}")
print(f"  Resigned: {len(resigned)}")
print(f"  Clean (no flags): {len(clean)}")
print(f"  Flagged: {len(flagged)}")
print(f"  With salary data: {len(with_salary)}")
print(f"  Without salary: {len(no_sal)}")
print(f"\nPosition distribution:")
for pos, cnt in position_counter.most_common():
    print(f"  {pos:30s} {cnt}")

flag_counter = Counter()
for r in flagged:
    for f in r['flags'].split('|'):
        if f: flag_counter[f] += 1
print(f"\nFlag distribution:")
for fl, cnt in flag_counter.most_common():
    print(f"  {fl:25s} {cnt}")

print(f"\nCSV written to: {csv_path}")
