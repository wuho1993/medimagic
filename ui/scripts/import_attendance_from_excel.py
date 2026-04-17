"""
Parse the attendance Excel (員工出勤資料.xlsx) and generate SQL to:
1. Create the monthly_attendance_records table (if not exists)
2. Insert per-employee, per-month attendance records linked to employees via employee_code
3. Upsert leave_balances for used leave days

Usage:
  python scripts/import_attendance_from_excel.py <path-to-excel>

Output:
  supabase/manual/20260416_attendance_import.sql
"""

import sys
import re
import json
from pathlib import Path
from collections import defaultdict

import openpyxl

# ---------------------------------------------------------------------------
# Column header normalization → canonical field names
# ---------------------------------------------------------------------------
HEADER_MAP = {
    'Staff Code': 'staff_code',
    'Itmes': 'name',
    'Items': 'name',
    # Attendance deduction
    '扣勤工': 'attendance_deduction',
    # Calendar / calculation days
    '計薪日數': 'calendar_days',
    '計薪 日數': 'calendar_days',
    '計菥日數': 'calendar_days',
    '計菥 日數': 'calendar_days',
    # Working days (month-prefixed)
    '上班': 'worked_days',
    # Days off
    '例假': 'off_days',
    # Statutory holidays
    '計算勞工假': 'statutory_holiday_days',
    # Birthday leave
    '生日假 (BL)': 'birthday_leave_days',
    '生日假(BL)': 'birthday_leave_days',
    '生 日假 (BL)': 'birthday_leave_days',
    # TB8
    'TB8': 'tb8_days',
    # Sick leave (paid)
    '病假\n(SL)': 'sick_leave_days',
    '病假': 'sick_leave_days',
    # Maternity leave
    '產假': 'maternity_leave_days',
    '產假\n': 'maternity_leave_days',
    # Reward leave
    '獎勵假(SB)': 'reward_leave_days',
    # Annual leave
    '年假(AL)': 'annual_leave_days',
    '年假': 'annual_leave_days',
    # Compassionate leave
    '恩恤假': 'compassionate_leave_days',
    # Sick no pay
    '病假\n(SL)(No Pay)': 'sick_no_pay_days',
    # No pay personal leave
    '事假\n(NPL)': 'no_pay_leave_days',
    '事假': 'no_pay_leave_days',
    # No pay statutory holiday
    'No Pay勞工假\n(NPSH)': 'no_pay_statutory_holiday_days',
    # OT / hours
    '上月剩餘鐘數 (HOUR)': 'prev_month_remaining_hours',
    '上月剩餘鐘數(HOUR)': 'prev_month_remaining_hours',
    '補鐘': 'makeup_hours',
    'OT': 'overtime_hours',
    '假期轉鐘(1天=8HOUR)': 'leave_to_hours_conversion',
    '假 期轉鐘(1天=8HOUR)': 'leave_to_hours_conversion',
    '累積OT時數 (Hours)': 'accumulated_ot_hours',
    '累積OT時數(Hours)': 'accumulated_ot_hours',
    '總日數': 'total_days',
    # Remarks
    'PS.': 'remarks',
    'PS': 'remarks',
    '備註': 'remarks',
}

# Branch section markers in column B
BRANCH_MARKERS = {
    'MKTOP', 'MKCY', 'TW', 'TMA', 'MOS', '大圍', 'CWB', 'OFFICE',
    '街霸 / 其他', '街霸/其他', '解痛館', 'PT',
}

NUMERIC_FIELDS = [
    'calendar_days', 'worked_days', 'off_days', 'statutory_holiday_days',
    'birthday_leave_days', 'tb8_days', 'sick_leave_days', 'maternity_leave_days',
    'reward_leave_days', 'annual_leave_days', 'compassionate_leave_days',
    'sick_no_pay_days', 'no_pay_leave_days', 'no_pay_statutory_holiday_days',
    'prev_month_remaining_hours', 'makeup_hours', 'overtime_hours',
    'leave_to_hours_conversion', 'accumulated_ot_hours', 'total_days',
]


def normalize_header(raw: str) -> str | None:
    """Try to match a raw header string to a canonical field name."""
    if not raw:
        return None
    raw_clean = raw.strip()
    # Direct match
    if raw_clean in HEADER_MAP:
        return HEADER_MAP[raw_clean]
    # Fuzzy: strip month prefixes like "3月" or "12月"
    stripped = re.sub(r'^\d{1,2}\s*月', '', raw_clean)
    if stripped in HEADER_MAP:
        return HEADER_MAP[stripped]
    # Try partial match
    for pattern, field in HEADER_MAP.items():
        if pattern in raw_clean or raw_clean in pattern:
            return field
    return None


def parse_year_month(sheet_name: str) -> str | None:
    """Convert sheet name to YYYY-MM format."""
    # "2026年3月" → "2026-03"
    m = re.match(r'(\d{4})年(\d{1,2})月', sheet_name)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    # "12月2024" → "2024-12"
    m = re.match(r'(\d{1,2})月(\d{4})', sheet_name)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    # "3月" → assume 2025 for months without year (Jan 2025 - Dec 2025 sheets)
    m = re.match(r'^(\d{1,2})月$', sheet_name)
    if m:
        return f"2025-{int(m.group(1)):02d}"
    return None


def is_staff_code(val: str | None) -> bool:
    if not val or not isinstance(val, str):
        return False
    return bool(re.match(r'^SF\d{3}$', val.strip()))


def is_branch_marker(val: str | None) -> bool:
    if not val or not isinstance(val, str):
        return False
    return val.strip() in BRANCH_MARKERS or val.strip().startswith('街霸')


def safe_numeric(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val.replace(',', ''))
        except ValueError:
            return None
    return None


def parse_sheet(ws, year_month: str) -> list[dict]:
    """Parse one worksheet into a list of attendance records."""
    # Step 1: build header mapping (col_letter → field_name)
    col_map = {}
    for cell in ws[1]:
        if cell.value:
            field = normalize_header(str(cell.value))
            if field:
                col_map[cell.column_letter] = field

    if 'staff_code' not in col_map.values():
        # Fallback: assume B is staff_code, C is name
        col_map['B'] = 'staff_code'
        col_map['C'] = 'name'

    records = []
    current_branch = None

    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for cell in ws[row_idx]:
            # Skip merged cells that don't have column_letter
            try:
                col_letter = cell.column_letter
            except AttributeError:
                continue
            field = col_map.get(col_letter)
            if field and cell.value is not None:
                row_data[field] = cell.value

        # Check for branch marker
        b_val = ws.cell(row=row_idx, column=2).value
        if b_val and isinstance(b_val, str) and is_branch_marker(b_val):
            current_branch = b_val.strip()
            continue

        # Skip empty rows or non-data rows
        staff_code = row_data.get('staff_code')
        if not is_staff_code(staff_code):
            continue

        # Build record
        rec = {
            'year_month': year_month,
            'employee_code': staff_code.strip(),
            'name': str(row_data.get('name', '')).strip(),
            'branch_section': current_branch,
        }
        for f in NUMERIC_FIELDS:
            rec[f] = safe_numeric(row_data.get(f))
        rec['attendance_deduction'] = bool(row_data.get('attendance_deduction'))
        rec['remarks'] = str(row_data.get('remarks', '')).strip() if row_data.get('remarks') else None

        records.append(rec)

    return records


def generate_sql(all_records: list[dict], output_path: Path):
    """Generate a single SQL file with table creation + data inserts + leave balance updates."""

    # Collect unique employee_codes
    emp_codes = sorted({r['employee_code'] for r in all_records})

    # Collect leave usage per employee per year for leave_balances upserts
    # leave types: annual, sick, birthday, maternity, reward, no_pay, compassionate, sick_no_pay, compensation (TB8)
    leave_usage = defaultdict(float)  # (emp_code, year, leave_type) → total days

    for r in all_records:
        year = int(r['year_month'].split('-')[0])
        ec = r['employee_code']
        for field, leave_type in [
            ('annual_leave_days', 'annual'),
            ('sick_leave_days', 'sick'),
            ('birthday_leave_days', 'birthday'),
            ('maternity_leave_days', 'maternity'),
            ('reward_leave_days', 'reward'),
            ('no_pay_leave_days', 'no_pay'),
            ('compassionate_leave_days', 'special'),
            ('sick_no_pay_days', 'sick_no_pay'),
            ('tb8_days', 'compensation'),
        ]:
            val = r.get(field)
            if val and val > 0:
                leave_usage[(ec, year, leave_type)] += val

    lines = []
    lines.append("-- Auto-generated attendance import from 員工出勤資料.xlsx")
    lines.append(f"-- Generated records: {len(all_records)}")
    lines.append(f"-- Unique employees: {len(emp_codes)}")
    lines.append(f"-- Employee codes: {', '.join(emp_codes)}")
    lines.append("")
    lines.append("begin;")
    lines.append("")

    # ---------------------------------------------------------------------------
    # 1. Create monthly_attendance_records table
    # ---------------------------------------------------------------------------
    lines.append("-- ============================================================")
    lines.append("-- 1. Create monthly_attendance_records table")
    lines.append("-- ============================================================")
    lines.append("""
create table if not exists public.monthly_attendance_records (
  id uuid primary key default gen_random_uuid(),
  employee_id uuid not null references public.employees(id) on delete cascade,
  year_month text not null,
  branch_section text,
  calendar_days numeric default 0,
  worked_days numeric default 0,
  off_days numeric default 0,
  statutory_holiday_days numeric default 0,
  birthday_leave_days numeric default 0,
  tb8_days numeric default 0,
  sick_leave_days numeric default 0,
  maternity_leave_days numeric default 0,
  reward_leave_days numeric default 0,
  annual_leave_days numeric default 0,
  compassionate_leave_days numeric default 0,
  sick_no_pay_days numeric default 0,
  no_pay_leave_days numeric default 0,
  no_pay_statutory_holiday_days numeric default 0,
  attendance_deduction boolean not null default false,
  prev_month_remaining_hours numeric default 0,
  makeup_hours numeric default 0,
  overtime_hours numeric default 0,
  leave_to_hours_conversion numeric default 0,
  accumulated_ot_hours numeric default 0,
  total_days numeric default 0,
  remarks text,
  created_at timestamptz not null default timezone('utc', now()),
  updated_at timestamptz not null default timezone('utc', now()),
  unique (employee_id, year_month)
);

comment on table public.monthly_attendance_records is 'Monthly attendance summary per employee, imported from attendance Excel workbook';
comment on column public.monthly_attendance_records.year_month is 'YYYY-MM format month identifier';
comment on column public.monthly_attendance_records.branch_section is 'Branch/section the employee was listed under in the source data';
comment on column public.monthly_attendance_records.calendar_days is '計薪日數 - calculation days for the month';
comment on column public.monthly_attendance_records.worked_days is '上班日數 - actual days worked';
comment on column public.monthly_attendance_records.off_days is '例假 (OFF) - regular days off';
comment on column public.monthly_attendance_records.statutory_holiday_days is '勞工假 (SH) - statutory holidays taken';
comment on column public.monthly_attendance_records.birthday_leave_days is '生日假 (BL) - birthday leave days';
comment on column public.monthly_attendance_records.tb8_days is 'TB8 days';
comment on column public.monthly_attendance_records.sick_leave_days is '有薪病假 (SL) - paid sick leave days';
comment on column public.monthly_attendance_records.maternity_leave_days is '產假 - maternity leave days';
comment on column public.monthly_attendance_records.reward_leave_days is '獎勵假 (SB) - reward/bonus leave days';
comment on column public.monthly_attendance_records.annual_leave_days is '年假 (AL) - annual leave days used';
comment on column public.monthly_attendance_records.compassionate_leave_days is '恩恤假 - compassionate leave days';
comment on column public.monthly_attendance_records.sick_no_pay_days is '無薪病假 (SL No Pay)';
comment on column public.monthly_attendance_records.no_pay_leave_days is '事假 (NPL) - no pay personal leave';
comment on column public.monthly_attendance_records.no_pay_statutory_holiday_days is 'No Pay 勞工假 (NPSH)';
comment on column public.monthly_attendance_records.attendance_deduction is '扣勤工 - whether attendance deduction applies';
comment on column public.monthly_attendance_records.prev_month_remaining_hours is '上月剩餘鐘數 - hours carried over from previous month';
comment on column public.monthly_attendance_records.makeup_hours is '補鐘 - makeup hours';
comment on column public.monthly_attendance_records.overtime_hours is 'OT - overtime hours';
comment on column public.monthly_attendance_records.leave_to_hours_conversion is '假期轉鐘 - leave days converted to hours (1 day = 8 hours)';
comment on column public.monthly_attendance_records.accumulated_ot_hours is '累積OT時數 - accumulated overtime hours balance';
comment on column public.monthly_attendance_records.total_days is '總日數 - total days accounted for the month';

create or replace function public.set_attendance_calendar_days()
returns trigger
language plpgsql
as $$
begin
    if new.year_month ~ '^\\d{4}-\\d{2}$' and coalesce(new.calendar_days, 0) <= 0 then
        new.calendar_days := extract(
            day from ((to_date(new.year_month || '-01', 'YYYY-MM-DD') + interval '1 month') - interval '1 day')
        );
    end if;

    return new;
end;
$$;

alter table public.monthly_attendance_records enable row level security;

drop policy if exists "Authenticated users can read attendance records" on public.monthly_attendance_records;
create policy "Authenticated users can read attendance records"
on public.monthly_attendance_records
for select to authenticated using (true);

drop policy if exists "Authenticated users can manage attendance records" on public.monthly_attendance_records;
create policy "Authenticated users can manage attendance records"
on public.monthly_attendance_records
for all to authenticated using (true) with check (true);

drop trigger if exists set_monthly_attendance_records_updated_at on public.monthly_attendance_records;
create trigger set_monthly_attendance_records_updated_at
before update on public.monthly_attendance_records
for each row execute function public.set_updated_at();

drop trigger if exists set_monthly_attendance_records_calendar_days on public.monthly_attendance_records;
create trigger set_monthly_attendance_records_calendar_days
before insert or update on public.monthly_attendance_records
for each row execute function public.set_attendance_calendar_days();
""")

    # ---------------------------------------------------------------------------
    # 2. Insert attendance records (only for employees that exist in DB)
    # ---------------------------------------------------------------------------
    lines.append("-- ============================================================")
    lines.append("-- 2. Insert attendance records (linked to employees table)")
    lines.append("-- ============================================================")
    lines.append("-- Only inserts for employee_codes that exist in the employees table.")
    lines.append("-- Missing codes are silently skipped via the subquery.")
    lines.append("")

    for r in all_records:
        vals = {
            'calendar_days': r.get('calendar_days') or 0,
            'worked_days': r.get('worked_days') or 0,
            'off_days': r.get('off_days') or 0,
            'statutory_holiday_days': r.get('statutory_holiday_days') or 0,
            'birthday_leave_days': r.get('birthday_leave_days') or 0,
            'tb8_days': r.get('tb8_days') or 0,
            'sick_leave_days': r.get('sick_leave_days') or 0,
            'maternity_leave_days': r.get('maternity_leave_days') or 0,
            'reward_leave_days': r.get('reward_leave_days') or 0,
            'annual_leave_days': r.get('annual_leave_days') or 0,
            'compassionate_leave_days': r.get('compassionate_leave_days') or 0,
            'sick_no_pay_days': r.get('sick_no_pay_days') or 0,
            'no_pay_leave_days': r.get('no_pay_leave_days') or 0,
            'no_pay_statutory_holiday_days': r.get('no_pay_statutory_holiday_days') or 0,
            'attendance_deduction': str(r.get('attendance_deduction', False)).lower(),
            'prev_month_remaining_hours': r.get('prev_month_remaining_hours') or 0,
            'makeup_hours': r.get('makeup_hours') or 0,
            'overtime_hours': r.get('overtime_hours') or 0,
            'leave_to_hours_conversion': r.get('leave_to_hours_conversion') or 0,
            'accumulated_ot_hours': r.get('accumulated_ot_hours') or 0,
            'total_days': r.get('total_days') or 0,
        }
        branch = f"'{r['branch_section']}'" if r.get('branch_section') else 'null'
        remarks_escaped = r['remarks'].replace("'", "''") if r.get('remarks') else None
        remarks_sql = f"'{remarks_escaped}'" if remarks_escaped else 'null'

        lines.append(f"insert into public.monthly_attendance_records (")
        lines.append(f"  employee_id, year_month, branch_section,")
        lines.append(f"  calendar_days, worked_days, off_days, statutory_holiday_days,")
        lines.append(f"  birthday_leave_days, tb8_days, sick_leave_days, maternity_leave_days,")
        lines.append(f"  reward_leave_days, annual_leave_days, compassionate_leave_days,")
        lines.append(f"  sick_no_pay_days, no_pay_leave_days, no_pay_statutory_holiday_days,")
        lines.append(f"  attendance_deduction, prev_month_remaining_hours, makeup_hours,")
        lines.append(f"  overtime_hours, leave_to_hours_conversion, accumulated_ot_hours, total_days,")
        lines.append(f"  remarks")
        lines.append(f")")
        lines.append(f"select")
        lines.append(f"  e.id, '{r['year_month']}', {branch},")
        lines.append(f"  {vals['calendar_days']}, {vals['worked_days']}, {vals['off_days']}, {vals['statutory_holiday_days']},")
        lines.append(f"  {vals['birthday_leave_days']}, {vals['tb8_days']}, {vals['sick_leave_days']}, {vals['maternity_leave_days']},")
        lines.append(f"  {vals['reward_leave_days']}, {vals['annual_leave_days']}, {vals['compassionate_leave_days']},")
        lines.append(f"  {vals['sick_no_pay_days']}, {vals['no_pay_leave_days']}, {vals['no_pay_statutory_holiday_days']},")
        lines.append(f"  {vals['attendance_deduction']}, {vals['prev_month_remaining_hours']}, {vals['makeup_hours']},")
        lines.append(f"  {vals['overtime_hours']}, {vals['leave_to_hours_conversion']}, {vals['accumulated_ot_hours']}, {vals['total_days']},")
        lines.append(f"  {remarks_sql}")
        lines.append(f"from public.employees e")
        lines.append(f"where e.employee_code = '{r['employee_code']}'")
        lines.append(f"on conflict (employee_id, year_month) do update set")
        lines.append(f"  branch_section = excluded.branch_section,")
        lines.append(f"  calendar_days = excluded.calendar_days,")
        lines.append(f"  worked_days = excluded.worked_days,")
        lines.append(f"  off_days = excluded.off_days,")
        lines.append(f"  statutory_holiday_days = excluded.statutory_holiday_days,")
        lines.append(f"  birthday_leave_days = excluded.birthday_leave_days,")
        lines.append(f"  tb8_days = excluded.tb8_days,")
        lines.append(f"  sick_leave_days = excluded.sick_leave_days,")
        lines.append(f"  maternity_leave_days = excluded.maternity_leave_days,")
        lines.append(f"  reward_leave_days = excluded.reward_leave_days,")
        lines.append(f"  annual_leave_days = excluded.annual_leave_days,")
        lines.append(f"  compassionate_leave_days = excluded.compassionate_leave_days,")
        lines.append(f"  sick_no_pay_days = excluded.sick_no_pay_days,")
        lines.append(f"  no_pay_leave_days = excluded.no_pay_leave_days,")
        lines.append(f"  no_pay_statutory_holiday_days = excluded.no_pay_statutory_holiday_days,")
        lines.append(f"  attendance_deduction = excluded.attendance_deduction,")
        lines.append(f"  prev_month_remaining_hours = excluded.prev_month_remaining_hours,")
        lines.append(f"  makeup_hours = excluded.makeup_hours,")
        lines.append(f"  overtime_hours = excluded.overtime_hours,")
        lines.append(f"  leave_to_hours_conversion = excluded.leave_to_hours_conversion,")
        lines.append(f"  accumulated_ot_hours = excluded.accumulated_ot_hours,")
        lines.append(f"  total_days = excluded.total_days,")
        lines.append(f"  remarks = excluded.remarks,")
        lines.append(f"  updated_at = timezone('utc', now());")
        lines.append("")

    # ---------------------------------------------------------------------------
    # 3. Upsert leave_balances from aggregated leave usage
    # ---------------------------------------------------------------------------
    lines.append("-- ============================================================")
    lines.append("-- 3. Update leave_balances with used_days from attendance data")
    lines.append("-- ============================================================")
    lines.append("")

    for (ec, year, leave_type), total_used in sorted(leave_usage.items()):
        lines.append(f"insert into public.leave_balances (employee_id, leave_type, year, used_days)")
        lines.append(f"select e.id, '{leave_type}'::leave_type, {year}, {total_used}")
        lines.append(f"from public.employees e where e.employee_code = '{ec}'")
        lines.append(f"on conflict (employee_id, leave_type, year) do update set")
        lines.append(f"  used_days = {total_used},")
        lines.append(f"  updated_at = timezone('utc', now());")
        lines.append("")

    lines.append("commit;")

    output_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f"✅ Generated SQL: {output_path}")
    print(f"   Records: {len(all_records)}")
    print(f"   Unique employees: {len(emp_codes)}")
    print(f"   Leave balance upserts: {len(leave_usage)}")


def main():
    if len(sys.argv) < 2:
        # Default path
        excel_path = Path('/Users/joecheung/Desktop/2026 Mar員工出勤資料.xlsx')
    else:
        excel_path = Path(sys.argv[1])

    if not excel_path.exists():
        print(f"❌ File not found: {excel_path}")
        sys.exit(1)

    print(f"📖 Reading: {excel_path.name}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    all_records = []
    skipped_sheets = []

    for sheet_name in wb.sheetnames:
        year_month = parse_year_month(sheet_name)
        if not year_month:
            skipped_sheets.append(sheet_name)
            continue

        ws = wb[sheet_name]
        records = parse_sheet(ws, year_month)
        print(f"  📋 {sheet_name} → {year_month}: {len(records)} employees")
        all_records.extend(records)

    if skipped_sheets:
        print(f"  ⚠️  Skipped sheets: {', '.join(skipped_sheets)}")

    # Generate SQL
    output_path = Path(__file__).parent.parent / 'supabase' / 'manual' / '20260416_attendance_import.sql'
    generate_sql(all_records, output_path)

    # Print summary of employee codes found in Excel vs staff import CSV
    excel_codes = sorted({r['employee_code'] for r in all_records})
    csv_path = Path(__file__).parent.parent / 'import' / 'full_staff_staging_20260415.csv'
    if csv_path.exists():
        import csv
        with open(csv_path) as f:
            csv_codes = {row['code'] for row in csv.DictReader(f)}
        missing = set(excel_codes) - csv_codes
        if missing:
            print(f"\n⚠️  Employee codes in Excel but NOT in staff CSV: {sorted(missing)}")
        else:
            print(f"\n✅ All {len(excel_codes)} employee codes found in staff CSV")


if __name__ == '__main__':
    main()
