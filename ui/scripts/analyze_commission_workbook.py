from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


KEYWORDS = [
    'redeem',
    'sales',
    'sgm',
    'bonus',
    'tl',
    '佣',
    '包佣',
    '街霸',
    '電話',
    '電話銷售',
    '人頭',
    '頭數',
    'job',
]

FORMULA_TOKENS = ['IF', 'IFS', 'VLOOKUP', 'XLOOKUP', 'SUM', 'ROUND', 'MAX', 'MIN']
EMPLOYEE_MARKERS = ['name', 'staff', 'employee', '員工', '姓名', '店舖', '職位', 'redeem', 'sales']


def normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ''


def sheet_preview_rows(ws, max_rows: int = 12, max_cols: int = 16) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True), start=1):
        values = [normalize_text(cell) for cell in row[:max_cols]]
        if any(values):
            rows.append({'row': row_index, 'values': [value for value in values if value]})
    return rows


def keyword_matches(ws, limit: int = 80) -> list[dict[str, str]]:
    matches: list[dict[str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            text = normalize_text(value)
            upper = text.upper()
            lower = text.lower()

            if text.startswith('='):
                if any(token in upper for token in FORMULA_TOKENS) or any(keyword in lower for keyword in KEYWORDS):
                    matches.append({'cell': cell.coordinate, 'value': text[:220]})
            elif any(keyword in lower for keyword in KEYWORDS):
                matches.append({'cell': cell.coordinate, 'value': text[:220]})

            if len(matches) >= limit:
                return matches
    return matches


def likely_header_rows(ws, scan_rows: int = 40) -> list[dict[str, object]]:
    headers: list[dict[str, object]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, scan_rows), values_only=True), start=1):
        values = [normalize_text(cell) for cell in row if normalize_text(cell)]
        if not values:
            continue
        joined = ' '.join(values).lower()
        if any(marker in joined for marker in EMPLOYEE_MARKERS):
            headers.append({'row': row_index, 'values': values[:20]})
    return headers[:8]


def main() -> int:
    if len(sys.argv) < 2:
        print('Usage: analyze_commission_workbook.py <xlsx-path>')
        return 1

    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    wb = load_workbook(workbook_path, data_only=False, read_only=True)

    result = {
        'file': str(workbook_path),
        'sheetNames': wb.sheetnames,
        'sheets': [],
    }

    for ws in wb.worksheets:
        result['sheets'].append(
            {
                'name': ws.title,
                'maxRow': ws.max_row,
                'maxColumn': ws.max_column,
                'previewRows': sheet_preview_rows(ws),
                'likelyHeaderRows': likely_header_rows(ws),
                'matches': keyword_matches(ws),
            }
        )

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())