#!/usr/bin/env python3
"""
Consolidate employee data from:
1. 員工資料.xlsx (personal info: 現職 + 離職 sheets)
2. 03-2026_MM(FY)_V6.xlsx (SALARY sheet: salary data + company type)

Outputs:
- consolidated_employees.txt (human-readable)
- 20260415_employee_import_v3.sql (DB import SQL)
"""

import openpyxl
import json
import re
from datetime import datetime, date

# ============================================================
# 1. PARSE EMPLOYEE INFO (員工資料.xlsx)
# ============================================================
wb_info = openpyxl.load_workbook('/Users/joecheung/Desktop/員工資料.xlsx', data_only=True)

def parse_date(val):
    """Parse various date formats to ISO string."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    # "1995年7月2日" format
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', s)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    # "19/1//2024" or "19/1/2024" 
    m = re.match(r'(\d{1,2})/(\d{1,2})/?/?(\d{4})', s)
    if m:
        return f'{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}'
    # "10月31日LAST DAY" → use 2024 context
    m = re.match(r'(\d{1,2})月(\d{1,2})日', s)
    if m:
        return f'2024-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
    # "20/11 LAST DAY"
    m = re.match(r'(\d{1,2})/(\d{1,2})\s', s)
    if m:
        return f'2024-{int(m.group(2)):02d}-{int(m.group(1)):02d}'
    # "3月29日"
    m = re.match(r'(\d{1,2})月(\d{1,2})日', s)
    if m:
        return f'2024-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
    return None

def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()

# --- Parse 現職 (Active) sheet ---
ws_active = wb_info['現職']
active_employees = {}
for row in range(2, ws_active.max_row + 1):
    code = safe_str(ws_active.cell(row=row, column=5).value)
    if not code or not code.startswith('SF'):
        continue
    branch = safe_str(ws_active.cell(row=row, column=2).value)
    position = safe_str(ws_active.cell(row=row, column=3).value)
    alias = safe_str(ws_active.cell(row=row, column=4).value)
    phone = safe_str(ws_active.cell(row=row, column=6).value)
    name_zh = safe_str(ws_active.cell(row=row, column=7).value)
    name_en = safe_str(ws_active.cell(row=row, column=8).value)
    dob = parse_date(ws_active.cell(row=row, column=9).value)
    id_number = safe_str(ws_active.cell(row=row, column=10).value)
    hire_date = parse_date(ws_active.cell(row=row, column=11).value)
    address = safe_str(ws_active.cell(row=row, column=12).value)
    
    active_employees[code] = {
        'code': code,
        'branch': branch,
        'position': position,
        'alias': alias,
        'phone': phone,
        'name_zh': name_zh,
        'name_en': name_en,
        'dob': dob,
        'id_number': id_number,
        'hire_date': hire_date,
        'address': address,
        'status': 'active',
        'end_date': None,
    }

# --- Parse 離職 (Resigned) sheet ---
ws_resigned = wb_info['離職']
resigned_employees = {}
for row in range(2, ws_resigned.max_row + 1):
    end_raw = ws_resigned.cell(row=row, column=1).value
    code = safe_str(ws_resigned.cell(row=row, column=5).value)
    if not code or not code.startswith('SF'):
        continue
    
    branch = safe_str(ws_resigned.cell(row=row, column=2).value)
    phone = safe_str(ws_resigned.cell(row=row, column=6).value)
    name_zh = safe_str(ws_resigned.cell(row=row, column=7).value)
    name_en = safe_str(ws_resigned.cell(row=row, column=8).value)
    dob = parse_date(ws_resigned.cell(row=row, column=9).value)
    id_number = safe_str(ws_resigned.cell(row=row, column=10).value)
    hire_date = parse_date(ws_resigned.cell(row=row, column=11).value)
    end_date = parse_date(end_raw)
    
    # Get position from column 3 if available
    position = safe_str(ws_resigned.cell(row=row, column=3).value)
    alias = safe_str(ws_resigned.cell(row=row, column=4).value)
    address_col = 12  # may or may not exist
    address = ''
    try:
        address = safe_str(ws_resigned.cell(row=row, column=14).value)
    except:
        pass
    
    resigned_employees[code] = {
        'code': code,
        'branch': branch,
        'position': position,
        'alias': alias,
        'phone': phone,
        'name_zh': name_zh,
        'name_en': name_en,
        'dob': dob,
        'id_number': id_number,
        'hire_date': hire_date,
        'address': address,
        'status': 'resigned',
        'end_date': end_date,
    }

print(f"Active employees from info sheet: {len(active_employees)}")
print(f"Resigned employees from info sheet: {len(resigned_employees)}")

# ============================================================
# 2. PARSE SALARY DATA (03-2026_MM(FY)_V6.xlsx)
# ============================================================
wb_salary = openpyxl.load_workbook('/Users/joecheung/Desktop/03-2026_MM(FY)_V6.xlsx', data_only=True)
ws_salary = wb_salary['SALARY']

salary_data = {}
for row in range(2, 200):
    code_raw = ws_salary.cell(row=row, column=2).value
    if code_raw is None:
        continue
    code = str(code_raw).strip()
    if code == 'N/A' or not code.startswith('SF'):
        continue
    
    basic_raw = ws_salary.cell(row=row, column=3).value
    attendance_raw = ws_salary.cell(row=row, column=4).value
    briefing_raw = ws_salary.cell(row=row, column=5).value
    booking_raw = ws_salary.cell(row=row, column=6).value
    transport_raw = ws_salary.cell(row=row, column=7).value
    company_raw = ws_salary.cell(row=row, column=15).value
    branch_raw = ws_salary.cell(row=row, column=14).value
    name_raw = ws_salary.cell(row=row, column=16).value
    
    # Determine salary type
    salary_type = 'monthly'
    base_salary = None
    
    if isinstance(basic_raw, (int, float)):
        base_salary = float(basic_raw)
    elif isinstance(basic_raw, str):
        if '/day' in basic_raw.lower():
            salary_type = 'daily'
            m = re.match(r'(\d+)', basic_raw)
            if m:
                base_salary = float(m.group(1))
        elif '/hr' in basic_raw.lower():
            salary_type = 'hourly'
            m = re.match(r'(\d+)', basic_raw)
            if m:
                base_salary = float(m.group(1))
        else:
            m = re.match(r'[\d.]+', basic_raw)
            if m:
                base_salary = float(m.group())
    
    # Parse bonus fields (may be N/A, '-', or numeric)
    def parse_num(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s in ('N/A', '-', ''):
            return None
        if s.startswith('/'):  # "/day" etc
            return None
        try:
            return float(s)
        except ValueError:
            return None
    
    attendance = parse_num(attendance_raw)
    briefing = parse_num(briefing_raw)
    booking = parse_num(booking_raw)
    transport = parse_num(transport_raw)
    
    # Company type
    company = None
    if company_raw:
        c = str(company_raw).strip().upper()
        if c == 'ASA' or c == 'Asa'.upper():
            company = 'ASA'
        elif c == 'ASAS':
            company = 'ASAS'
    
    salary_data[code] = {
        'code': code,
        'salary_type': salary_type,
        'base_salary': base_salary,
        'attendance': attendance,
        'briefing': briefing,
        'booking': booking,
        'transport': transport,
        'company': company,
        'branch_salary': str(branch_raw).strip() if branch_raw else None,
        'name_salary': str(name_raw).strip() if name_raw else None,
    }

print(f"Salary records: {len(salary_data)}")

# ============================================================
# 3. CONSOLIDATE
# ============================================================
# All SF codes from both sources
all_codes = set()
all_codes.update(active_employees.keys())
all_codes.update(resigned_employees.keys())
all_codes.update(salary_data.keys())

# Build unified list
consolidated = []
for code in sorted(all_codes):
    info = active_employees.get(code, resigned_employees.get(code, None))
    sal = salary_data.get(code, None)
    
    rec = {
        'code': code,
        'name_zh': '',
        'name_en': '',
        'alias': '',
        'phone': '',
        'address': '',
        'dob': None,
        'id_number': '',
        'hire_date': None,
        'end_date': None,
        'branch': '',
        'position': '',
        'status': 'active',
        'company_type': 'ASA',
        'salary_type': 'monthly',
        'base_salary': None,
        'attendance': None,
        'briefing': None,
        'booking': None,
        'transport': None,
        'employment_type': '全職',
    }
    
    # Fill from employee info
    if info:
        rec['name_zh'] = info['name_zh']
        rec['name_en'] = info['name_en']
        rec['alias'] = info['alias']
        rec['phone'] = info['phone']
        rec['address'] = info['address']
        rec['dob'] = info['dob']
        rec['id_number'] = info['id_number']
        rec['hire_date'] = info['hire_date']
        rec['end_date'] = info['end_date']
        rec['branch'] = info['branch']
        rec['position'] = info['position']
        rec['status'] = info['status']
    
    # Fill/override from salary
    if sal:
        rec['company_type'] = sal['company'] or rec['company_type']
        rec['salary_type'] = sal['salary_type']
        rec['base_salary'] = sal['base_salary']
        rec['attendance'] = sal['attendance']
        rec['briefing'] = sal['briefing']
        rec['booking'] = sal['booking']
        rec['transport'] = sal['transport']
        # If no info, use salary sheet name/branch
        if not rec['alias'] and sal['name_salary']:
            rec['alias'] = sal['name_salary']
        if not rec['branch'] and sal['branch_salary']:
            rec['branch'] = sal['branch_salary']
    
    # Determine if active but only in resigned sheet
    if code in resigned_employees and code not in active_employees and code in salary_data:
        # Has salary = likely still active, override
        rec['status'] = 'active'
        rec['end_date'] = None
    
    # Part-time check
    if rec['alias'] and ('PT' in rec['alias'] or '(PT)' in rec['alias']):
        rec['employment_type'] = '兼職'
    
    consolidated.append(rec)

print(f"Total consolidated records: {len(consolidated)}")
active_count = sum(1 for r in consolidated if r['status'] == 'active')
resigned_count = sum(1 for r in consolidated if r['status'] == 'resigned')
print(f"  Active: {active_count}, Resigned: {resigned_count}")

# ============================================================
# 4. IDENTIFY STREET PROMOTERS from image data
# ============================================================
# From image 1, bottom rows show: Dan Wu (16500+2000), Lilly (16000+2000), Wu Ann (18000)
# These have 街霸 label in the sheet
street_promoter_info = {
    # These will be added from image data manually since they appear at bottom of SALARY sheet
}

# ============================================================
# 5. GENERATE TXT FILE
# ============================================================
output_lines = []
output_lines.append("=" * 140)
output_lines.append("MEDI MAGIC HRMS - CONSOLIDATED EMPLOYEE DATA")
output_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
output_lines.append(f"Total: {len(consolidated)} employees (Active: {active_count}, Resigned: {resigned_count})")
output_lines.append("=" * 140)
output_lines.append("")

# Active employees
output_lines.append("─" * 140)
output_lines.append("ACTIVE EMPLOYEES")
output_lines.append("─" * 140)
header = f"{'Code':<8} {'Alias':<15} {'Name(ZH)':<12} {'Name(EN)':<25} {'Branch':<8} {'Position':<22} {'Company':<6} {'SalType':<10} {'Base':>10} {'Atten':>8} {'Brief':>8} {'Book':>8} {'Trans':>8}"
output_lines.append(header)
output_lines.append("-" * 140)

for r in consolidated:
    if r['status'] != 'active':
        continue
    base_str = f"{r['base_salary']:.0f}" if r['base_salary'] else ''
    att_str = f"{r['attendance']:.0f}" if r['attendance'] else ''
    brf_str = f"{r['briefing']:.0f}" if r['briefing'] else ''
    bk_str = f"{r['booking']:.0f}" if r['booking'] else ''
    tr_str = f"{r['transport']:.0f}" if r['transport'] else ''
    
    line = f"{r['code']:<8} {r['alias']:<15} {r['name_zh']:<12} {r['name_en']:<25} {r['branch']:<8} {r['position']:<22} {r['company_type']:<6} {r['salary_type']:<10} {base_str:>10} {att_str:>8} {brf_str:>8} {bk_str:>8} {tr_str:>8}"
    output_lines.append(line)

output_lines.append("")
output_lines.append("─" * 140)
output_lines.append("RESIGNED EMPLOYEES")
output_lines.append("─" * 140)
header2 = f"{'Code':<8} {'Alias':<15} {'Name(ZH)':<12} {'Name(EN)':<25} {'Branch':<8} {'Position':<22} {'Hire Date':<12} {'End Date':<12}"
output_lines.append(header2)
output_lines.append("-" * 140)

for r in consolidated:
    if r['status'] != 'resigned':
        continue
    hire_str = r['hire_date'] or ''
    end_str = r['end_date'] or ''
    line = f"{r['code']:<8} {r['alias']:<15} {r['name_zh']:<12} {r['name_en']:<25} {r['branch']:<8} {r['position']:<22} {hire_str:<12} {end_str:<12}"
    output_lines.append(line)

output_lines.append("")
output_lines.append("─" * 140)
output_lines.append("EMPLOYEES WITH SALARY BUT NO PERSONAL INFO (need data entry)")
output_lines.append("─" * 140)
for r in consolidated:
    if r['status'] == 'active' and not r['name_zh'] and not r['name_en']:
        output_lines.append(f"  {r['code']} - {r['alias']} (Branch: {r['branch']}, Company: {r['company_type']})")

txt_path = '/Users/joecheung/Kojin Ai Solution/Medi Magic HRMS/ui/supabase/manual/consolidated_employees.txt'
with open(txt_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))

print(f"\nTXT file written to: {txt_path}")

# ============================================================
# 6. EXPORT JSON for SQL generation
# ============================================================
json_path = '/Users/joecheung/Kojin Ai Solution/Medi Magic HRMS/ui/supabase/manual/consolidated_employees.json'
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(consolidated, f, ensure_ascii=False, indent=2, default=str)

print(f"JSON file written to: {json_path}")
