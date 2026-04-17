from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path('/Users/joecheung/Kojin Ai Solution/02-2026_MM(FY)_V8.xlsx')


def safe_text(value: object) -> str:
    if value is None:
        return ''
    return str(value).strip()


def inspect_role_patterns(workbook) -> None:
    ws = workbook['Contract List']
    role_counter: Counter[str] = Counter()
    special_rows: list[tuple[int, str, str, str, str, str, str, str]] = []

    for row in range(1, ws.max_row + 1):
        staff_no = safe_text(ws.cell(row, 1).value)
        nickname = safe_text(ws.cell(row, 2).value)
        role = safe_text(ws.cell(row, 4).value)
        package_floor = safe_text(ws.cell(row, 21).value)
        target = safe_text(ws.cell(row, 22).value)
        package_text = safe_text(ws.cell(row, 23).value)
        period_or_limit = safe_text(ws.cell(row, 24).value)
        commission_note = safe_text(ws.cell(row, 25).value)

        if nickname and role and role != '職位':
            role_counter[role] += 1

        interesting = [package_floor, target, package_text, period_or_limit, commission_note]
        if nickname and role and role != '職位' and any(
            value and value not in {'N/A', '公司另行通知'} for value in interesting
        ):
            special_rows.append(
                (row, staff_no, nickname, role, package_floor, target, package_text, commission_note)
            )

    print('=== ROLE COUNTS ===')
    for role, count in role_counter.most_common():
        print(f'{role}\t{count}')

    print('\n=== CONTRACT LIST SPECIAL COMMISSION ROWS ===')
    for item in special_rows:
        print('\t'.join(str(value) for value in item))


def inspect_salary_commission_relation(workbook) -> None:
    salary = workbook['SALARY']
    commission = workbook['COMMISSION']

    commission_by_code: dict[str, dict[str, str]] = {}
    for row in range(2, commission.max_row + 1):
        code = safe_text(commission.cell(row, 2).value)
        name = safe_text(commission.cell(row, 3).value)
        if not code:
            continue
        commission_by_code[code] = {
            'row': str(row),
            'name': name,
            'job': safe_text(commission.cell(row, 4).value),
            'redeem': safe_text(commission.cell(row, 5).value),
            'sales': safe_text(commission.cell(row, 7).value),
            'sales_formula': safe_text(commission.cell(row, 9).value),
            'sgm': safe_text(commission.cell(row, 10).value),
        }

    print('\n=== SALARY <-> COMMISSION MATCHES ===')
    for row in range(25, 160):
        code = safe_text(salary.cell(row, 2).value)
        name = safe_text(salary.cell(row, 15).value)
        if not code or code not in commission_by_code:
            continue
        info = commission_by_code[code]
        salary_job = safe_text(salary.cell(row, 21).value)
        salary_redeem = safe_text(salary.cell(row, 22).value)
        salary_redeem_formula = safe_text(salary.cell(row, 23).value)
        salary_sales = safe_text(salary.cell(row, 24).value)
        salary_sales_tl_formula = safe_text(salary.cell(row, 25).value)
        salary_sgm = safe_text(salary.cell(row, 26).value)
        salary_sgm_tl = safe_text(salary.cell(row, 27).value)
        print(
            '\t'.join(
                [
                    f'SALARY_ROW={row}',
                    f'CODE={code}',
                    f'SALARY_NAME={name}',
                    f'COMM_ROW={info["row"]}',
                    f'COMM_NAME={info["name"]}',
                    f'JOB={salary_job}',
                    f'COMM_JOB={info["job"]}',
                    f'REDEEM={salary_redeem}',
                    f'COMM_REDEEM={info["redeem"]}',
                    f'SALES={salary_sales}',
                    f'COMM_SALES={info["sales"]}',
                    f'REDEEM_FORMULA={salary_redeem_formula}',
                    f'SALES_TL_FORMULA={salary_sales_tl_formula}',
                    f'SGM={salary_sgm}',
                    f'SGM_TL={salary_sgm_tl}',
                ]
            )
        )


def inspect_formula_patterns(workbook) -> None:
    salary = workbook['SALARY']
    sample_rows = [27, 28, 29, 30, 31, 32, 33, 34, 35, 39, 40, 41, 42, 43]

    print('\n=== SALARY FORMULA PATTERNS ===')
    for row in sample_rows:
        code = safe_text(salary.cell(row, 2).value)
        name = safe_text(salary.cell(row, 15).value)
        print(
            '\t'.join(
                [
                    f'ROW={row}',
                    f'CODE={code}',
                    f'NAME={name}',
                    f'V_REDEEM={safe_text(salary.cell(row, 22).value)}',
                    f'W_REDEEM_FORMULA={safe_text(salary.cell(row, 23).value)}',
                    f'X_SALES={safe_text(salary.cell(row, 24).value)}',
                    f'Y_SALES_TL={safe_text(salary.cell(row, 25).value)}',
                    f'AA_SGM_TL={safe_text(salary.cell(row, 27).value)}',
                    f'AK_PACKAGE={safe_text(salary.cell(row, 37).value)}',
                    f'AN_TOTAL={safe_text(salary.cell(row, 40).value)}',
                    f'AP_7TH={safe_text(salary.cell(row, 42).value)}',
                    f'AR_MPF={safe_text(salary.cell(row, 44).value)}',
                    f'AT_20TH={safe_text(salary.cell(row, 46).value)}',
                    f'AV_REASON={safe_text(salary.cell(row, 48).value)}',
                ]
            )
        )


def main() -> int:
    workbook = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    inspect_role_patterns(workbook)
    inspect_salary_commission_relation(workbook)
    inspect_formula_patterns(workbook)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())