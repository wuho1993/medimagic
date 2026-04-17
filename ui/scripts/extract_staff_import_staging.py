from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
WORKBOOK_PATH = ROOT / '02-2026_MM(FY)_V8.xlsx'
OUTPUT_DIR = ROOT / 'Medi Magic HRMS' / 'ui' / 'import'
STAGING_PATH = OUTPUT_DIR / 'staff_import_staging_20260415.csv'
REVIEW_PATH = OUTPUT_DIR / 'staff_import_review_20260415.csv'


BRANCH_MAP = {
    'OFFICE': 'OFFICE',
    'OFFICE ': 'OFFICE',
    'OFFICE/': 'OFFICE',
    'OFFICE/ ': 'OFFICE',
    'OFfice': 'OFFICE',
    'Office': 'OFFICE',
    'TOP': 'MKTOP',
    'MKTOP': 'MKTOP',
    'TM': 'TMA',
    'TAIWAI': 'TAIWAI',
    'TaiWai': 'TAIWAI',
    'TW': 'TW',
    'MOS': 'MOS',
    'MKCY': 'MKCY',
    'CBA': 'CBA',
}

POSITION_MAP = {
    '治療師': 'BEAUTICIAN',
    '養生師': 'BEAUTICIAN',
    '接待員': 'CONSULTANT',
    '顧問': 'CONSULTANT',
    '店長': 'CONSULTANT',
    '經理': 'CONSULTANT',
    '會計經理': 'CONSULTANT',
    '會計': 'CONSULTANT',
    '文員': 'CONSULTANT',
    '客戶服務主管': 'CONSULTANT',
    '培訓老師': 'CONSULTANT',
    '電話銷售員': 'CONSULTANT',
    '老闆': 'CONSULTANT',
}

COMPANY_MAP = {
    'A': 'ASA',
    'ASA': 'ASA',
    'ASAS': 'ASAS',
    '自': 'ASA',
}


@dataclass
class StaffRow:
    employee_code: str
    alias: str
    branch_raw: str
    branch_normalized: str | None
    position_raw: str
    position_normalized: str | None
    name_en: str
    name_zh: str
    identity_number: str
    company_raw: str
    company_normalized: str | None
    probation_months: int | None
    hire_date: str | None
    employment_status: str
    employment_type: str
    gender: str
    base_salary: str
    attendance_bonus_amount: str
    briefing_bonus: str
    booking_bonus: str
    transport_allowance: str
    pay_day_primary: str
    pay_day_secondary: str
    review_flags: str


def normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ''


def normalize_date(value: object) -> str | None:
    if value is None or value == '':
      return None
    if isinstance(value, datetime):
      return value.strftime('%Y-%m-%d')
    text = normalize_text(value)
    text = text.replace('年', '-').replace('月', '-').replace('日', '')
    text = text.replace('/', '-')
    parts = [part for part in text.split('-') if part]
    if len(parts) == 3 and all(part.isdigit() for part in parts):
      year, month, day = parts
      return f'{year.zfill(4)}-{month.zfill(2)}-{day.zfill(2)}'
    return None


def normalize_probation(value: object) -> int | None:
    text = normalize_text(value)
    digits = ''.join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None


def normalize_branch(value: str) -> tuple[str | None, list[str]]:
    text = normalize_text(value)
    flags: list[str] = []
    if not text:
        flags.append('missing_branch')
        return None, flags
    normalized = BRANCH_MAP.get(text)
    if normalized:
        return normalized, flags
    if '/' in text:
        flags.append('multi_branch_review')
        return None, flags
    if text.upper() in BRANCH_MAP:
        return BRANCH_MAP[text.upper()], flags
    flags.append('unknown_branch')
    return None, flags


def normalize_position(value: str) -> tuple[str | None, list[str]]:
    text = normalize_text(value).replace(' ', '')
    flags: list[str] = []
    if not text:
        flags.append('missing_position')
        return None, flags
    normalized = POSITION_MAP.get(text)
    if normalized:
        return normalized, flags
    flags.append('unknown_position')
    return None, flags


def normalize_company(value: str) -> tuple[str | None, list[str]]:
    text = normalize_text(value).upper()
    flags: list[str] = []
    if not text:
        flags.append('missing_company')
        return None, flags
    normalized = COMPANY_MAP.get(text)
    if normalized:
        return normalized, flags
    flags.append('unknown_company')
    return None, flags


def infer_status(code: str) -> str:
    return 'resigned' if 'resigned' in code.lower() else 'active'


def normalize_code(code: str) -> tuple[str, list[str]]:
    text = normalize_text(code)
    flags: list[str] = []
    if ' ' in text:
        text = text.split(' ')[0]
        flags.append('code_trimmed')
    return text, flags


def iter_staff_rows() -> Iterable[StaffRow]:
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=True)
    ws = wb['Contract List']
    for row in ws.iter_rows(values_only=True):
        raw_code = normalize_text(row[0] if len(row) > 0 else '')
        if not raw_code.startswith('SF'):
            continue
        employee_code, code_flags = normalize_code(raw_code)
        alias = normalize_text(row[1] if len(row) > 1 else '')
        branch_raw = normalize_text(row[2] if len(row) > 2 else '')
        position_raw = normalize_text(row[3] if len(row) > 3 else '')
        name_en = normalize_text(row[4] if len(row) > 4 else '')
        name_zh = normalize_text(row[5] if len(row) > 5 else '')
        identity_number = normalize_text(row[6] if len(row) > 6 else '')
        company_raw = normalize_text(row[9] if len(row) > 9 else '')
        probation_months = normalize_probation(row[10] if len(row) > 10 else None)
        hire_date = normalize_date(row[12] if len(row) > 12 else None)
        base_salary = normalize_text(row[14] if len(row) > 14 else '')
        attendance_bonus_amount = normalize_text(row[15] if len(row) > 15 else '')
        briefing_bonus = normalize_text(row[16] if len(row) > 16 else '')
        booking_bonus = normalize_text(row[17] if len(row) > 17 else '')
        transport_allowance = normalize_text(row[18] if len(row) > 18 else '')
        pay_day_primary = normalize_text(row[26] if len(row) > 26 else '')
        pay_day_secondary = normalize_text(row[27] if len(row) > 27 else '')

        branch_normalized, branch_flags = normalize_branch(branch_raw)
        position_normalized, position_flags = normalize_position(position_raw)
        company_normalized, company_flags = normalize_company(company_raw)

        flags = code_flags + branch_flags + position_flags + company_flags
        if not identity_number or identity_number == 'N/A':
            flags.append('missing_identity_number')
        if not hire_date:
            flags.append('missing_hire_date')
        if not name_en:
            flags.append('missing_name_en')
        if not name_zh:
            flags.append('missing_name_zh')

        yield StaffRow(
            employee_code=employee_code,
            alias=alias,
            branch_raw=branch_raw,
            branch_normalized=branch_normalized,
            position_raw=position_raw,
            position_normalized=position_normalized,
            name_en=name_en,
            name_zh=name_zh,
            identity_number=identity_number,
            company_raw=company_raw,
            company_normalized=company_normalized,
            probation_months=probation_months,
            hire_date=hire_date,
            employment_status=infer_status(raw_code),
            employment_type='full_time',
            gender='other',
            base_salary=base_salary,
            attendance_bonus_amount=attendance_bonus_amount,
            briefing_bonus=briefing_bonus,
            booking_bonus=booking_bonus,
            transport_allowance=transport_allowance,
            pay_day_primary=pay_day_primary,
            pay_day_secondary=pay_day_secondary,
            review_flags='|'.join(flags),
        )


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = list(iter_staff_rows())
    fieldnames = list(StaffRow.__dataclass_fields__.keys())

    with STAGING_PATH.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.__dict__)

    with REVIEW_PATH.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            if row.review_flags:
                writer.writerow(row.__dict__)

    print(f'Wrote {len(rows)} staged staff rows to {STAGING_PATH}')
    print(f'Wrote {sum(1 for row in rows if row.review_flags)} review rows to {REVIEW_PATH}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())